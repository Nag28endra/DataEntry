from flask import Flask, render_template, request
from openpyxl import Workbook
from openpyxl import load_workbook
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    date = request.form['date']
    course = request.form['course']
    activity = request.form['activity']

    # Create a new workbook and sheet
    wb = Workbook()
    sheet = wb.active

    # Check if the Excel file already exists, and append to it if it does
    if os.path.exists('data.xlsx'):
        wb = load_workbook('data.xlsx')
        sheet = wb.active
    else:
        # Add headers if this is the first time data is being added
        sheet['A1'] = 'Name of the Student'
        sheet['B1'] = 'Date'
        sheet['C1'] = 'Course Learned'
        sheet['D1'] = 'Activity Done'

    # Add data to the Excel sheet
    row = (name, date, course, activity)
    sheet.append(row)

    # Save the workbook
    wb.save('data.xlsx')

    return 'Data saved successfully!'

if __name__ == '__main__':
    app.run(debug=True)
